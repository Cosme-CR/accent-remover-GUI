import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import unicodedata
import xml.etree.ElementTree as ET
import threading
import os
import csv as csv_module

from openpyxl import load_workbook
import pandas as pd

try:
    import xlwt
except ImportError:
    xlwt = None


# ============================================================
# REMOVER ACENTOS
# ============================================================

def remover_acentos(texto):
    """
    Remove acentos de textos.

    Exemplo:
        João -> Joao
        São Paulo -> Sao Paulo
        Informação -> Informacao
    """

    if texto is None or not isinstance(texto, str):
        return texto

    texto = unicodedata.normalize("NFD", texto)

    return "".join(
        caractere for caractere in texto
        if unicodedata.category(caractere) != "Mn"
    )


def remover_acentos_dataframe(dataframe):
    """
    Aplica remover_acentos em todas as células de texto de um DataFrame,
    preservando números, datas e outros tipos.
    """

    funcao = lambda valor: remover_acentos(valor) if isinstance(valor, str) else valor

    # pandas >= 2.1 usa .map ; versões antigas usam .applymap
    if hasattr(dataframe, "map"):
        try:
            return dataframe.map(funcao)
        except TypeError:
            pass

    return dataframe.applymap(funcao)


# ============================================================
# XML
# ============================================================

def processar_elemento_xml(elemento):
    """
    Percorre todo o XML e remove acentos de textos, tails e atributos.
    """

    if elemento.text:
        elemento.text = remover_acentos(elemento.text)

    if elemento.tail:
        elemento.tail = remover_acentos(elemento.tail)

    for nome, valor in elemento.attrib.items():
        elemento.attrib[nome] = remover_acentos(valor)

    for filho in elemento:
        processar_elemento_xml(filho)


def remover_acentos_xml(arquivo_entrada, arquivo_saida):
    """
    Processa arquivo XML.
    """

    parser = ET.XMLParser(encoding="utf-8")
    tree = ET.parse(arquivo_entrada, parser=parser)
    root = tree.getroot()

    processar_elemento_xml(root)

    tree.write(arquivo_saida, encoding="utf-8", xml_declaration=True)

    return arquivo_saida


# ============================================================
# XLSX -> XLSX (caminho rápido, preserva fórmulas e formatação)
# ============================================================

def remover_acentos_xlsx(arquivo_entrada, arquivo_saida):
    """
    Processa XLSX diretamente com openpyxl, célula a célula.

    Isso evita reconstruir o arquivo inteiro através do pandas e reduz
    bastante a possibilidade de o Excel considerar o arquivo corrompido.
    Fórmulas e formatação são preservadas.
    """

    workbook = load_workbook(arquivo_entrada, data_only=False)

    for planilha in workbook.worksheets:
        for linha in planilha.iter_rows():
            for celula in linha:
                valor = celula.value

                if isinstance(valor, str) and not valor.startswith("="):
                    novo_valor = remover_acentos(valor)

                    if novo_valor != valor:
                        celula.value = novo_valor

    workbook.save(arquivo_saida)

    return arquivo_saida


# ============================================================
# LEITURA DE TABELAS (XLS, XLSX, CSV) EM DATAFRAMES
# ============================================================

def detectar_csv(arquivo_entrada):
    """
    Tenta identificar a codificação e o separador de um arquivo CSV.
    """

    codificacoes = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]

    amostra = None
    codificacao_usada = codificacoes[-1]

    for codificacao in codificacoes:
        try:
            with open(arquivo_entrada, "r", encoding=codificacao) as arquivo:
                amostra = arquivo.read(8192)
            codificacao_usada = codificacao
            break
        except (UnicodeDecodeError, LookupError):
            continue

    separador = ","

    if amostra:
        try:
            dialeto = csv_module.Sniffer().sniff(amostra, delimiters=[",", ";", "\t", "|"])
            separador = dialeto.delimiter
        except csv_module.Error:
            # Fallback simples: escolhe o separador mais frequente na amostra
            candidatos = [",", ";", "\t", "|"]
            separador = max(candidatos, key=lambda caractere: amostra.count(caractere))

    return codificacao_usada, separador


def assinatura_arquivo(arquivo_entrada, tamanho=4096):
    """
    Lê os primeiros bytes do arquivo para identificar seu formato real,
    independente da extensão informada.
    """

    with open(arquivo_entrada, "rb") as arquivo:
        return arquivo.read(tamanho)


def tipo_real_arquivo(arquivo_entrada):
    """
    Muitos sistemas (governo, ERPs, sistemas legados) geram arquivos
    chamados ".xls" que na verdade são XML (formato "Excel XML
    Spreadsheet" / SpreadsheetML) ou até HTML. Isso causa o erro
    "Expected BOF record" quando se tenta abrir com xlrd.

    Esta função olha o conteúdo real do arquivo e identifica o formato,
    para que a leitura correta seja usada independente da extensão.
    """

    dados = assinatura_arquivo(arquivo_entrada)

    if dados.startswith(b"PK"):
        return "xlsx"

    if dados[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"

    texto = dados.decode("utf-8-sig", errors="ignore").lstrip()
    texto_lower = texto.lower()

    if texto_lower.startswith("<?xml") or texto_lower.startswith("<workbook"):
        return "spreadsheetml"

    if "<html" in texto_lower or "<table" in texto_lower:
        return "html"

    return "desconhecido"


def ler_spreadsheetml(arquivo_entrada):
    """
    Lê o formato "Excel XML Spreadsheet" (SpreadsheetML), muito comum em
    exportações de sistemas legados que salvam arquivos XML com a
    extensão .xls.
    """

    tree = ET.parse(arquivo_entrada)
    root = tree.getroot()

    if root.tag.startswith("{"):
        uri = root.tag[1:].split("}")[0]
    else:
        uri = "urn:schemas-microsoft-com:office:spreadsheet"

    ns = {"ss": uri}

    planilhas = {}

    for indice_planilha, worksheet in enumerate(root.findall("ss:Worksheet", ns)):
        nome_planilha = worksheet.attrib.get(
            "{%s}Name" % uri, f"Sheet{indice_planilha + 1}"
        )

        tabela = worksheet.find("ss:Table", ns)
        linhas_dados = []

        if tabela is not None:
            for linha_xml in tabela.findall("ss:Row", ns):
                indice_coluna = 0
                valores_por_coluna = {}

                for celula in linha_xml.findall("ss:Cell", ns):
                    indice_attr = celula.attrib.get("{%s}Index" % uri)

                    if indice_attr:
                        indice_coluna = int(indice_attr) - 1

                    elemento_dado = celula.find("ss:Data", ns)
                    valor = elemento_dado.text if elemento_dado is not None else None

                    valores_por_coluna[indice_coluna] = valor
                    indice_coluna += 1

                if valores_por_coluna:
                    coluna_maxima = max(valores_por_coluna.keys())
                    linha_lista = [
                        valores_por_coluna.get(coluna) for coluna in range(coluna_maxima + 1)
                    ]
                else:
                    linha_lista = []

                linhas_dados.append(linha_lista)

        largura_maxima = max((len(linha) for linha in linhas_dados), default=0)
        linhas_normalizadas = [
            linha + [None] * (largura_maxima - len(linha)) for linha in linhas_dados
        ]

        planilhas[str(nome_planilha)[:31]] = pd.DataFrame(linhas_normalizadas)

    if not planilhas:
        raise ValueError(
            "O arquivo é um XML, mas não foi possível encontrar nenhuma "
            "planilha reconhecível dentro dele."
        )

    return planilhas


def carregar_tabela(arquivo_entrada, extensao):
    """
    Carrega XLS, XLSX ou CSV em um dicionário {nome_planilha: DataFrame}.
    Retorna também metadados (codificação/separador) quando aplicável.

    Para XLS/XLSX, o conteúdo real do arquivo é verificado primeiro, pois
    é comum receber arquivos com a extensão errada (por exemplo, um XML
    ou HTML salvo como ".xls").
    """

    if extensao == ".csv":
        codificacao, separador = detectar_csv(arquivo_entrada)

        dataframe = pd.read_csv(
            arquivo_entrada,
            header=None,
            sep=separador,
            engine="python",
            encoding=codificacao,
            keep_default_na=False,
            na_values=[],
        )

        return {"Sheet1": dataframe}, {"codificacao": codificacao, "separador": separador}

    if extensao in (".xls", ".xlsx"):
        tipo_real = tipo_real_arquivo(arquivo_entrada)

        if tipo_real == "xlsx":
            planilhas = pd.read_excel(
                arquivo_entrada, sheet_name=None, header=None, engine="openpyxl"
            )
            return planilhas, None

        if tipo_real == "xls":
            planilhas = pd.read_excel(
                arquivo_entrada, sheet_name=None, header=None, engine="xlrd"
            )
            return planilhas, None

        if tipo_real == "spreadsheetml":
            return ler_spreadsheetml(arquivo_entrada), None

        if tipo_real == "html":
            tabelas = pd.read_html(arquivo_entrada, header=None)
            return (
                {f"Sheet{indice + 1}": tabela for indice, tabela in enumerate(tabelas)},
                None,
            )

        amostra = assinatura_arquivo(arquivo_entrada, 60)
        amostra_texto = amostra.decode("utf-8", errors="replace")

        raise ValueError(
            f"O arquivo tem extensão {extensao}, mas seu conteúdo não é um "
            "Excel válido nem uma planilha XML/HTML reconhecida.\n\n"
            f"Início do arquivo: {amostra_texto!r}"
        )

    raise ValueError("Formato de tabela não suportado.")


def salvar_como_xlsx(planilhas, arquivo_saida):
    """
    Salva um dicionário de DataFrames como XLSX (uma aba por planilha).
    """

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        for nome_planilha, dataframe in planilhas.items():
            nome_planilha = str(nome_planilha)[:31] or "Planilha1"

            dataframe.to_excel(
                writer, sheet_name=nome_planilha, header=False, index=False
            )

    return arquivo_saida


def salvar_como_xls(planilhas, arquivo_saida):
    """
    Salva um dicionário de DataFrames como XLS (formato antigo, via xlwt).

    Limitações do formato XLS: no máximo 65.536 linhas e 256 colunas
    por planilha, e sem fórmulas complexas ou formatação avançada.
    """

    if xlwt is None:
        raise ValueError(
            "A biblioteca 'xlwt' não está instalada.\n\n"
            "Instale com: pip install xlwt"
        )

    workbook = xlwt.Workbook(encoding="utf-8")
    planilhas_gravadas = 0
    linhas_cortadas = False

    for nome_planilha, dataframe in planilhas.items():
        nome_planilha = str(nome_planilha)[:31] or "Planilha1"

        aba = workbook.add_sheet(nome_planilha)

        linhas = dataframe.values.tolist()

        if len(linhas) > 65536:
            linhas = linhas[:65536]
            linhas_cortadas = True

        for indice_linha, linha in enumerate(linhas):
            for indice_coluna, valor in enumerate(linha[:256]):
                if pd.isna(valor):
                    valor = ""

                aba.write(indice_linha, indice_coluna, valor)

        planilhas_gravadas += 1

    workbook.save(arquivo_saida)

    return arquivo_saida, linhas_cortadas


def salvar_como_csv(planilhas, arquivo_saida, info_csv=None):
    """
    Salva a primeira planilha do dicionário como CSV.
    Se houver mais de uma planilha, as demais são ignoradas (CSV não
    suporta múltiplas abas).
    """

    primeira_planilha = next(iter(planilhas.values()))

    codificacao = "utf-8-sig"
    separador = ","

    if info_csv:
        codificacao = info_csv.get("codificacao", codificacao)
        separador = info_csv.get("separador", separador)

    primeira_planilha.to_csv(
        arquivo_saida,
        header=False,
        index=False,
        encoding=codificacao,
        sep=separador,
    )

    return arquivo_saida, len(planilhas) > 1


# ============================================================
# PROCESSAMENTO AUTOMÁTICO
# ============================================================

FORMATOS_TABULARES = (".xls", ".xlsx", ".csv")


def processar_arquivo(arquivo_entrada, arquivo_saida):
    """
    Identifica automaticamente os tipos de entrada/saída e processa o
    arquivo, removendo acentos de todo o conteúdo textual.

    Retorna (arquivo_final, aviso_ou_None).
    """

    extensao_entrada = os.path.splitext(arquivo_entrada)[1].lower()
    extensao_saida = os.path.splitext(arquivo_saida)[1].lower()

    # --------------------------------------------------------
    # XML
    # --------------------------------------------------------

    if extensao_entrada == ".xml":
        if extensao_saida != ".xml":
            raise ValueError("Arquivos XML devem ser salvos como .xml")

        return remover_acentos_xml(arquivo_entrada, arquivo_saida), None

    # --------------------------------------------------------
    # XLS / XLSX / CSV
    # --------------------------------------------------------

    if extensao_entrada in FORMATOS_TABULARES:
        if extensao_saida not in (".xlsx", ".xls", ".csv"):
            raise ValueError(
                "Para arquivos XLS, XLSX ou CSV, escolha uma saída "
                "em XLSX, XLS ou CSV."
            )

        # Caminho rápido: preserva fórmulas e formatação originais
        # (só é seguro quando o arquivo realmente é um .xlsx válido;
        # arquivos "falsos" .xlsx, como XML/HTML renomeados, caem no
        # caminho normal abaixo, via carregar_tabela)
        if (
            extensao_entrada == ".xlsx"
            and extensao_saida == ".xlsx"
            and tipo_real_arquivo(arquivo_entrada) == "xlsx"
        ):
            return remover_acentos_xlsx(arquivo_entrada, arquivo_saida), None

        planilhas, info_csv = carregar_tabela(arquivo_entrada, extensao_entrada)
        planilhas = {
            nome: remover_acentos_dataframe(dataframe)
            for nome, dataframe in planilhas.items()
        }

        if extensao_saida == ".xlsx":
            return salvar_como_xlsx(planilhas, arquivo_saida), None

        if extensao_saida == ".xls":
            arquivo_final, linhas_cortadas = salvar_como_xls(planilhas, arquivo_saida)

            aviso = None
            if linhas_cortadas:
                aviso = (
                    "O formato XLS suporta no máximo 65.536 linhas por "
                    "planilha. Linhas excedentes foram descartadas."
                )

            return arquivo_final, aviso

        arquivo_final, havia_varias_planilhas = salvar_como_csv(
            planilhas, arquivo_saida, info_csv
        )

        aviso = None
        if havia_varias_planilhas:
            aviso = (
                "O arquivo de origem possuía mais de uma planilha.\n"
                "Como CSV suporta apenas uma tabela, somente a primeira "
                "planilha foi salva."
            )

        return arquivo_final, aviso

    # --------------------------------------------------------
    # FORMATO INVÁLIDO
    # --------------------------------------------------------

    raise ValueError(
        "Formato não suportado.\n\nFormatos aceitos:\n• XML\n• XLS\n• XLSX\n• CSV"
    )


# ============================================================
# APARÊNCIA
# ============================================================

CORES = {
    "fundo": "#000000",
    "cartao": "#141414",
    "primaria": "#4C7BFF",
    "primaria_hover": "#6E92FF",
    "secundaria": "#1F1F1F",
    "secundaria_hover": "#2A2A2A",
    "texto": "#F5F6FA",
    "texto_suave": "#9AA0AC",
    "borda": "#2A2A2E",
    "sucesso": "#2ECC71",
    "erro": "#FF5C5C",
}

FONTE_BASE = "Segoe UI"


# ============================================================
# APLICAÇÃO
# ============================================================

class Aplicacao:

    def __init__(self, root):
        self.root = root
        self.root.title("Removedor de Acentos")
        self.root.geometry("760x560")
        self.root.minsize(960, 760)
        self.root.configure(bg=CORES["fundo"])

        self.entrada_var = tk.StringVar()
        self.saida_var = tk.StringVar()
        self.formato_saida_var = tk.StringVar()

        self.configurar_estilos()
        self.criar_interface()

    # ========================================================
    # ESTILOS
    # ========================================================

    def configurar_estilos(self):
        estilo = ttk.Style(self.root)

        try:
            estilo.theme_use("clam")
        except tk.TclError:
            pass

        estilo.configure(
            "TEntry",
            fieldbackground=CORES["secundaria"],
            foreground=CORES["texto"],
            insertcolor=CORES["texto"],
            bordercolor=CORES["borda"],
            lightcolor=CORES["borda"],
            darkcolor=CORES["borda"],
            padding=8,
        )
        estilo.map(
            "TEntry",
            fieldbackground=[("readonly", CORES["secundaria"])],
        )

        estilo.configure(
            "TCombobox",
            fieldbackground=CORES["secundaria"],
            background=CORES["secundaria"],
            foreground=CORES["texto"],
            arrowcolor=CORES["texto"],
            bordercolor=CORES["borda"],
            padding=6,
        )
        estilo.map(
            "TCombobox",
            fieldbackground=[("readonly", CORES["secundaria"])],
            foreground=[("readonly", CORES["texto"])],
        )
        self.root.option_add("*TCombobox*Listbox*Background", CORES["secundaria"])
        self.root.option_add("*TCombobox*Listbox*Foreground", CORES["texto"])
        self.root.option_add("*TCombobox*Listbox*selectBackground", CORES["primaria"])
        self.root.option_add("*TCombobox*Listbox*selectForeground", "#FFFFFF")

        # Botões secundários: bem visíveis sobre o fundo preto, com borda clara
        estilo.configure(
            "Secundario.TButton",
            font=(FONTE_BASE, 10, "bold"),
            padding=(16, 10),
            background=CORES["secundaria"],
            foreground=CORES["texto"],
            bordercolor=CORES["primaria"],
            borderwidth=1.5,
            relief="flat",
        )
        estilo.map(
            "Secundario.TButton",
            background=[("active", CORES["secundaria_hover"])],
            bordercolor=[("active", CORES["primaria_hover"])],
        )

        # Botão principal: cor cheia e chamativa
        estilo.configure(
            "Primario.TButton",
            font=(FONTE_BASE, 13, "bold"),
            padding=(22, 14),
            background=CORES["primaria"],
            foreground="#FFFFFF",
            borderwidth=0,
            relief="flat",
        )
        estilo.map(
            "Primario.TButton",
            background=[
                ("disabled", "#33406B"),
                ("active", CORES["primaria_hover"]),
            ],
            foreground=[("disabled", "#8A93B3")],
        )

        estilo.configure(
            "Fino.Horizontal.TProgressbar",
            troughcolor=CORES["secundaria"],
            background=CORES["primaria"],
            thickness=8,
            borderwidth=0,
        )

    # ========================================================
    # INTERFACE
    # ========================================================

    def criar_interface(self):
        # ---------------- Cabeçalho ----------------
        cabecalho = tk.Frame(self.root, bg=CORES["fundo"], height=110)
        cabecalho.pack(fill="x")
        cabecalho.pack_propagate(False)

        tk.Label(
            cabecalho,
            text="Removedor de Acentos",
            font=(FONTE_BASE, 24, "bold"),
            bg=CORES["fundo"],
            fg=CORES["primaria"],
        ).pack(pady=(26, 2))

        tk.Label(
            cabecalho,
            text="Suporta arquivos XML, XLS, XLSX e CSV",
            font=(FONTE_BASE, 10),
            bg=CORES["fundo"],
            fg=CORES["texto_suave"],
        ).pack()

        # ---------------- Cartão principal ----------------
        cartao = tk.Frame(self.root, bg=CORES["cartao"], highlightbackground=CORES["borda"], highlightthickness=1)
        cartao.pack(fill="both", expand=True, padx=30, pady=25)

        conteudo = tk.Frame(cartao, bg=CORES["cartao"])
        conteudo.pack(fill="both", expand=True, padx=28, pady=26)
        conteudo.columnconfigure(0, weight=1)

        # ---- Arquivo de entrada ----
        self.criar_rotulo(conteudo, "Arquivo de entrada", row=0)

        linha_entrada = tk.Frame(conteudo, bg=CORES["cartao"])
        linha_entrada.grid(row=1, column=0, sticky="ew", pady=(6, 4))
        linha_entrada.columnconfigure(0, weight=1)

        self.entrada_entry = ttk.Entry(linha_entrada, textvariable=self.entrada_var)
        self.entrada_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10), ipady=4)

        ttk.Button(
            linha_entrada,
            text="Selecionar arquivo",
            style="Secundario.TButton",
            command=self.selecionar_entrada,
        ).grid(row=0, column=1)

        self.dica_entrada = tk.Label(
            conteudo,
            text="Formatos aceitos: .xml  .xls  .xlsx  .csv  (saída também em .xls)",
            font=(FONTE_BASE, 9),
            bg=CORES["cartao"],
            fg=CORES["texto_suave"],
        )
        self.dica_entrada.grid(row=2, column=0, sticky="w", pady=(0, 22))

        # ---- Formato de saída ----
        self.criar_rotulo(conteudo, "Formato de saída", row=3)

        self.combo_formato = ttk.Combobox(
            conteudo,
            textvariable=self.formato_saida_var,
            state="readonly",
            values=[],
            width=20,
        )
        self.combo_formato.grid(row=4, column=0, sticky="w", pady=(6, 22))
        self.combo_formato.bind("<<ComboboxSelected>>", self.ao_trocar_formato_saida)

        # ---- Arquivo de saída ----
        self.criar_rotulo(conteudo, "Arquivo de saída", row=5)

        linha_saida = tk.Frame(conteudo, bg=CORES["cartao"])
        linha_saida.grid(row=6, column=0, sticky="ew", pady=(6, 4))
        linha_saida.columnconfigure(0, weight=1)

        self.saida_entry = ttk.Entry(linha_saida, textvariable=self.saida_var)
        self.saida_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10), ipady=4)

        ttk.Button(
            linha_saida,
            text="Salvar como",
            style="Secundario.TButton",
            command=self.selecionar_saida,
        ).grid(row=0, column=1)

        # ---- Status ----
        self.status_var = tk.StringVar(value="Aguardando seleção do arquivo de entrada...")

        self.status = tk.Label(
            conteudo,
            textvariable=self.status_var,
            font=(FONTE_BASE, 10),
            bg=CORES["cartao"],
            fg=CORES["texto_suave"],
            anchor="w",
            justify="left",
            wraplength=620,
        )
        self.status.grid(row=7, column=0, sticky="ew", pady=(20, 8))

        # ---- Progresso ----
        self.progresso = ttk.Progressbar(
            conteudo, mode="indeterminate", style="Fino.Horizontal.TProgressbar"
        )
        self.progresso.grid(row=8, column=0, sticky="ew", pady=(0, 20))

        # ---- Botão processar ----
        self.botao_processar = ttk.Button(
            conteudo,
            text="PROCESSAR ARQUIVO",
            style="Primario.TButton",
            command=self.iniciar_processamento,
            state="disabled",
        )
        self.botao_processar.grid(row=9, column=0, sticky="ew")

    def criar_rotulo(self, mestre, texto, row):
        tk.Label(
            mestre,
            text=texto,
            font=(FONTE_BASE, 10, "bold"),
            bg=CORES["cartao"],
            fg=CORES["texto"],
        ).grid(row=row, column=0, sticky="w")

    # ========================================================
    # SELECIONAR ENTRADA
    # ========================================================

    def selecionar_entrada(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo",
            filetypes=[
                ("Arquivos suportados", "*.xml *.xls *.xlsx *.csv"),
                ("Arquivo XML", "*.xml"),
                ("Excel", "*.xls *.xlsx"),
                ("CSV", "*.csv"),
                ("Todos os arquivos", "*.*"),
            ],
        )

        if not arquivo:
            return

        self.entrada_var.set(arquivo)

        extensao = os.path.splitext(arquivo)[1].lower()

        if extensao == ".xml":
            self.status_var.set("Arquivo XML selecionado.")
            opcoes = ["XML"]

        elif extensao == ".xls":
            self.status_var.set("Arquivo XLS selecionado.")
            opcoes = ["XLSX", "XLS", "CSV"]

        elif extensao == ".xlsx":
            self.status_var.set("Arquivo XLSX selecionado.")
            opcoes = ["XLSX", "XLS", "CSV"]

        elif extensao == ".csv":
            self.status_var.set("Arquivo CSV selecionado.")
            opcoes = ["CSV", "XLSX", "XLS"]

        else:
            self.status_var.set("Formato não suportado.")
            self.botao_processar.config(state="disabled")
            self.combo_formato.config(values=[])
            self.formato_saida_var.set("")
            return

        self.combo_formato.config(values=opcoes)
        self.formato_saida_var.set(opcoes[0])
        self.botao_processar.config(state="normal")

        self.sugerir_saida(arquivo, opcoes[0])

    # ========================================================
    # TROCAR FORMATO DE SAÍDA
    # ========================================================

    def ao_trocar_formato_saida(self, evento=None):
        arquivo_entrada = self.entrada_var.get().strip()

        if arquivo_entrada:
            self.sugerir_saida(arquivo_entrada, self.formato_saida_var.get())

    def sugerir_saida(self, arquivo_entrada, formato_saida):
        nome = os.path.splitext(os.path.basename(arquivo_entrada))[0]
        pasta = os.path.dirname(arquivo_entrada)

        extensao_saida = "." + formato_saida.lower()

        caminho_saida = os.path.join(pasta, f"{nome}_sem_acentos{extensao_saida}")

        self.saida_var.set(caminho_saida)

    # ========================================================
    # SELECIONAR SAÍDA
    # ========================================================

    def selecionar_saida(self):
        formato_saida = self.formato_saida_var.get() or "XLSX"
        extensao_saida = "." + formato_saida.lower()

        tipos = {
            "XML": [("Arquivo XML", "*.xml")],
            "XLSX": [("Excel XLSX", "*.xlsx")],
            "XLS": [("Excel XLS", "*.xls")],
            "CSV": [("Arquivo CSV", "*.csv")],
        }

        arquivo = filedialog.asksaveasfilename(
            title="Escolher arquivo de saída",
            defaultextension=extensao_saida,
            filetypes=tipos.get(formato_saida, [("Todos os arquivos", "*.*")]),
        )

        if arquivo:
            self.saida_var.set(arquivo)
            self.status_var.set("Local do arquivo de saída selecionado.")

    # ========================================================
    # INICIAR PROCESSAMENTO
    # ========================================================

    def iniciar_processamento(self):
        arquivo_entrada = self.entrada_var.get().strip()
        arquivo_saida = self.saida_var.get().strip()

        if not arquivo_entrada:
            messagebox.showwarning("Atenção", "Selecione o arquivo de entrada.")
            return

        if not os.path.isfile(arquivo_entrada):
            messagebox.showerror("Erro", "O arquivo de entrada não existe.")
            return

        if not arquivo_saida:
            messagebox.showwarning("Atenção", "Selecione onde o arquivo de saída será salvo.")
            return

        # Garante que a extensão de saída bate com o formato escolhido
        extensao_esperada = "." + (self.formato_saida_var.get() or "XLSX").lower()

        if not arquivo_saida.lower().endswith(extensao_esperada):
            arquivo_saida = os.path.splitext(arquivo_saida)[0] + extensao_esperada
            self.saida_var.set(arquivo_saida)

        if os.path.abspath(arquivo_entrada) == os.path.abspath(arquivo_saida):
            messagebox.showerror(
                "Erro",
                "O arquivo de entrada e o arquivo de saída não podem ser o mesmo arquivo.",
            )
            return

        pasta_saida = os.path.dirname(os.path.abspath(arquivo_saida))

        if not os.path.isdir(pasta_saida):
            try:
                os.makedirs(pasta_saida, exist_ok=True)
            except Exception as erro:
                messagebox.showerror(
                    "Erro", f"Não foi possível criar a pasta de saída.\n\n{erro}"
                )
                return

        if os.path.exists(arquivo_saida):
            resposta = messagebox.askyesno(
                "Arquivo já existe",
                "O arquivo de saída já existe.\n\nDeseja substituí-lo?",
            )

            if not resposta:
                return

        self.botao_processar.config(state="disabled")
        self.combo_formato.config(state="disabled")
        self.progresso.start(10)
        self.status_var.set("Processando arquivo...")

        thread = threading.Thread(
            target=self.processar,
            args=(arquivo_entrada, arquivo_saida),
            daemon=True,
        )
        thread.start()

    # ========================================================
    # PROCESSAR
    # ========================================================

    def processar(self, arquivo_entrada, arquivo_saida):
        try:
            arquivo_final, aviso = processar_arquivo(arquivo_entrada, arquivo_saida)
            self.root.after(0, self.processamento_sucesso, arquivo_final, aviso)

        except ET.ParseError as erro:
            self.root.after(
                0,
                self.processamento_erro,
                f"O arquivo selecionado não é um XML válido.\n\n{erro}",
            )

        except PermissionError as erro:
            self.root.after(
                0,
                self.processamento_erro,
                (
                    "Não foi possível salvar o arquivo.\n\n"
                    "Verifique se o arquivo de saída está aberto no Excel "
                    f"ou em outro programa.\n\n{erro}"
                ),
            )

        except Exception as erro:
            self.root.after(0, self.processamento_erro, str(erro))

    # ========================================================
    # SUCESSO
    # ========================================================

    def processamento_sucesso(self, arquivo_saida, aviso=None):
        self.progresso.stop()
        self.botao_processar.config(state="normal")
        self.combo_formato.config(state="readonly")

        self.status_var.set("Arquivo processado com sucesso!")

        mensagem = f"O arquivo foi processado com sucesso!\n\nArquivo salvo em:\n{arquivo_saida}"

        if aviso:
            mensagem += f"\n\nAviso:\n{aviso}"

        messagebox.showinfo("Concluído", mensagem)

    # ========================================================
    # ERRO
    # ========================================================

    def processamento_erro(self, mensagem):
        self.progresso.stop()
        self.botao_processar.config(state="normal")
        self.combo_formato.config(state="readonly")

        self.status_var.set("Erro durante o processamento.")

        messagebox.showerror("Erro", f"Não foi possível processar o arquivo.\n\n{mensagem}")


# ============================================================
# INICIAR PROGRAMA
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = Aplicacao(root)
    root.mainloop()